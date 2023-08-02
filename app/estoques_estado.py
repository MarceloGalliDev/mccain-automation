# flake8: noqa
import os
import logging
import openpyxl
import pandas as pd
from ftplib import FTP
from datetime import datetime
from dotenv import load_dotenv
from config import get_db_engine, log_data

load_dotenv()
log_data()


def estoques():
    FTP_CONFIG = {
        'server-ftp': os.getenv('SERVER-FTP'),
        'user-ftp': os.getenv('USER-FTP'),
        'password-ftp': os.getenv('PASSWORD-FTP'),
        'path_produto': os.getenv('PATH-PRODUTO'),
        'path_clientes': os.getenv('PATH-CLIENTES'),
        'path_clientes_sp': os.getenv('PATH-CLIENTES-SP'),
        'path_clientes_pr': os.getenv('PATH-CLIENTES-PR'),
        'path_estoque': os.getenv('PATH-ESTOQUE'),
        'path_estoque_sp': os.getenv('PATH-ESTOQUE-SP'),
        'path_estoque_pr': os.getenv('PATH-ESTOQUE-PR'),
        'path_vendas': os.getenv('PATH-VENDAS'),
        'path_vendas_SP': os.getenv('PATH-VENDAS-SP'),
        'path_vendas_PR': os.getenv('PATH-VENDAS-PR'),
    }

    unid_codigos = [['001','003'], '002']

    conn = get_db_engine()
    ftp = FTP_CONFIG

    for unid_codigo in unid_codigos:
        if isinstance(unid_codigo, list):
            unid_values = ",".join([f"'{code}'" for code in unid_codigo])
        else:
            unid_values = f"'{unid_codigo}'"
        query = (f"""
            (
                SELECT
                    prun.prun_estoque1 AS estoque,
                    (prun.prun_estoque1 * prod.prod_pesoliq) AS qtde,
                    prun.prun_unid_codigo AS unidade,
                    prun.prun_ativo as tipo,
                    prun.prun_prod_codigo AS prod_codigo,
                    prod.prod_codbarras AS cod_barras,
                    prod.prod_marca AS marca,
                    prod.prod_codigo AS cod_prod
                FROM produn AS prun
                LEFT JOIN produtos AS prod ON prun.prun_prod_codigo = prod.prod_codigo
                WHERE prun.prun_bloqueado = 'N'
                AND prun.prun_unid_codigo IN ({unid_values})
                AND prun.prun_ativo = 'S'
                AND prun.prun_estoque1 > 0
                AND prod.prod_marca IN ('MCCAIN','MCCAIN RETAIL')
            )
        """)

        df = pd.read_sql_query(query, conn)
        
        grouped_df = df.groupby('cod_prod').agg({'qtde': 'sum'}).reset_index()

        wb = openpyxl.Workbook()
        ws = wb.active

        dataAtualEstoque = datetime.now().strftime("%Y-%m-%d")
        ws['A1'] = ('Code')
        ws['B1'] = ('Quantity')
        ws['C1'] = ('Stock Date')
        ws['D1'] = ('Expiration Date')
        for index, row in grouped_df.iterrows():
            code = row["cod_prod"]
            quantity = row["qtde"]
            stockDate = dataAtualEstoque
            expirationDate = ''

            ws.cell(row=index+2, column=1).value = (f'{code:.0f}')
            ws.cell(row=index+2, column=2).value = (f'{quantity:.2f}')
            ws.cell(row=index+2, column=3).value = (f'{stockDate}')
            ws.cell(row=index+2, column=4).value = (f'{expirationDate}')

        if unid_codigo == ['001','003']:
            unid_codigo = 'PR'
        else:
            unid_codigo = 'SP'
        dataAtual = datetime.now().strftime("%Y-%m-%d")
        nomeArquivo = (f'ESTOQUEDUSNEI{unid_codigo}{dataAtual}')
        ws.title = dataAtual
        diretorio = f'C:/Users/Windows/Documents/Python/mccain-automation/app/data/{dataAtual}'
        if not os.path.exists(diretorio):
            os.mkdir(diretorio, exist_ok=True)
        local_arquivo = os.path.join(
            f'C:/Users/Windows/Documents/Python/mccain-automation/app/data/{dataAtual}/{nomeArquivo}.xlsx')

        wb.save(local_arquivo)


    with FTP(FTP_CONFIG['server-ftp']) as ftp:
        ftp.login(user=FTP_CONFIG['user-ftp'], passwd=FTP_CONFIG['password-ftp'])

        remote_dir_path_pr = os.path.join(FTP_CONFIG['path_estoque_pr'])
        remote_dir_path_sp = os.path.join(FTP_CONFIG['path_estoque_sp'])

        # try:
        #     ftp.mkd(remote_dir_path)
        #     print(f'Diretório {remote_dir_path} criado!')
        # except Exception as e:
        #     print('Não foi possível criar a pasta, pode ser que já exista!')

        for arquivos_data in os.listdir(diretorio):
            if 'ESTOQUEDUSNEIPR' in arquivos_data:
                file_path = os.path.join(diretorio, arquivos_data)

                if os.path.isfile(file_path):
                    with open(local_arquivo, 'rb') as local_file:
                        remote_path = os.path.join(remote_dir_path_pr, arquivos_data)
                        ftp.storbinary(f"STOR {remote_path}", local_file)
                logging.info(
                    f"Arquivo {os.path.basename(arquivos_data)} upload FTP server concluído com sucesso!")
                
            if 'ESTOQUEDUSNEISP' in arquivos_data:
                file_path = os.path.join(diretorio, arquivos_data)

                if os.path.isfile(file_path):
                    with open(local_arquivo, 'rb') as local_file:
                        remote_path = os.path.join(remote_dir_path_sp, arquivos_data)
                        ftp.storbinary(f"STOR {remote_path}", local_file)
                logging.info(
                    f"Arquivo {os.path.basename(arquivos_data)} upload FTP server concluído com sucesso!")


if __name__ == "__main__":
    estoques()