# flake8: noqa
import os
import logging
import openpyxl
import pandas as pd
from ftplib import FTP
from datetime import datetime
from dotenv import load_dotenv
from config import get_db_engine, log_data

load_dotenv()
log_data()


def vendas_estado():
    FTP_CONFIG = {
        'server-ftp': os.getenv('SERVER-FTP'),
        'user-ftp': os.getenv('USER-FTP'),
        'password-ftp': os.getenv('PASSWORD-FTP'),
        'path_produto': os.getenv('PATH-PRODUTO'),
        'path_clientes': os.getenv('PATH-CLIENTES'),
        'path_clientes_sp': os.getenv('PATH-CLIENTES-SP'),
        'path_clientes_pr': os.getenv('PATH-CLIENTES-PR'),
        'path_estoque': os.getenv('PATH-ESTOQUE'),
        'path_estoque_sp': os.getenv('PATH-ESTOQUE-SP'),
        'path_estoque_pr': os.getenv('PATH-ESTOQUE-PR'),
        'path_vendas': os.getenv('PATH-VENDAS'),
        'path_vendas_sp': os.getenv('PATH-VENDAS-SP'),
        'path_vendas_pr': os.getenv('PATH-VENDAS-PR'),
    }


    def vendas_query(table_name, conn):
        query = (f"""
            (
                SELECT
                    mprd.mprd_dcto_codigo AS doc_cod, 
                    mprd.mprd_transacao AS transacao,
                    clie.clie_cnpjcpf AS cnpj_cpf,
                    clie.clie_codigo AS cod_clie,
                    mprd.mprd_datamvto AS data,
                    mprd.mprd_numerodcto AS nfe,
                    prod.prod_codbarras AS cod_barras,
                    prod.prod_codigo AS cod_prod,
                    (mprd.mprd_qtde * prod.prod_pesoliq) AS quantity,
                    mprd.mprd_valor AS amount,
                    mprc.mprc_vend_codigo AS cod_vend,
                    mprc.mprc_uf AS estado,
                    prod.prod_marca AS marca,
                    SUBSTRING(clie.clie_cepres, 1,5) ||'-'|| SUBSTRING(clie.clie_cepres, 6,3) AS cep
                FROM {table_name} AS mprd 
                LEFT JOIN movprodc AS mprc ON mprd.mprd_operacao = mprc.mprc_operacao
                LEFT JOIN produtos AS prod ON mprd.mprd_prod_codigo = prod.prod_codigo
                LEFT JOIN clientes AS clie ON mprc.mprc_codentidade = clie.clie_codigo
                WHERE mprd_status = 'N' 
                AND prod.prod_marca IN ('MCCAIN','MCCAIN RETAIL')
                AND mprd.mprd_dcto_codigo IN ('6666','6668','7335','7337','7338','7339','7260','7263','7262','7268','7264','7269', '7267', '7319', '7318')
                AND mprc.mprc_uf = 'PR'
            )  
        """)
                # AND mprd.mprd_datamvto > CURRENT_DATE - INTERVAL '8 DAYS'
        return pd.read_sql_query(query, conn)


    conn = get_db_engine()
    ftp = FTP_CONFIG
    
    
    tables = [
        'movprodd0123', 
        'movprodd0223', 
        'movprodd0323', 
        'movprodd0423', 
        'movprodd0523', 
        'movprodd0623', 
        'movprodd0723', 
        'movprodd0823', 
        # 'movprodd0923', 
        # 'movprodd1023', 
        # 'movprodd1123', 
        # 'movprodd1223'
    ]
    
    for table in tables:
        data_frame = vendas_query(table, conn)
        print(f'Obtidos {data_frame.shape[0]} registros da tabela {table}.')
    
        wb = openpyxl.Workbook()
        ws = wb.active
           
        ws['A1'] = ('systemId')
        ws['B1'] = ('Code')
        ws['C1'] = ('Quantity')
        ws['D1'] = ('Amount')
        ws['E1'] = ('Sale Date')
        ws['F1'] = ('Transaction ID')
        for index, row in data_frame.iterrows():
            systemId = row["cod_clie"]
            code = row["cod_prod"]
            doc_cod = row["doc_cod"]
            quantity = row["quantity"]

            if doc_cod in ['7260', '7263', '7262', '7268', '7264', '7269', '7267', '7319', '7318']:
                quantity = -quantity

            amount = str(row["amount"]).replace(',', '.')
            amount2 = float(amount)
            data = row["data"].strftime("%Y-%m-%d")
            transactionId = "1" + row["nfe"].zfill(6)

            ws.cell(row=index+2, column=1).value = (f'{systemId:.0f}')
            ws.cell(row=index+2, column=2).value = (f'{code:.0f}')
            ws.cell(row=index+2, column=3).value = (f'{quantity:.2f}')
            ws.cell(row=index+2, column=4).value = (f'{amount2:.2f}')
            ws.cell(row=index+2, column=5).value = (f'{data}')
            ws.cell(row=index+2, column=6).value = (f'{transactionId}')

            dataAtual = datetime.now().strftime("%Y-%m-%d")
            nomeArquivo = (f'VENDASDUSNEIPR{table}{dataAtual}')
            ws.title = dataAtual
            diretorio = f'C:/Users/Windows/Documents/Python/mccain-automation/app/data/{dataAtual}'
            if not os.path.exists(diretorio):
                os.mkdir(diretorio)
            local_arquivo = os.path.join(
                f'C:/Users/Windows/Documents/Python/mccain-automation/app/data/{dataAtual}/{nomeArquivo}.xlsx')

            wb.save(local_arquivo)
    print("Processamento concluído!")


if __name__ == "__main__":
    vendas_estado()
  