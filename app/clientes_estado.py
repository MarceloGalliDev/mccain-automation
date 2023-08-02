# flake8: noqa
import os
import logging
import openpyxl
import pandas as pd
from ftplib import FTP
from datetime import datetime
from dotenv import load_dotenv
from config import get_db_engine, log_data

load_dotenv()
log_data()

def clientes():
    FTP_CONFIG = {
        'server-ftp': os.getenv('SERVER-FTP'),
        'user-ftp': os.getenv('USER-FTP'),
        'password-ftp': os.getenv('PASSWORD-FTP'),
        'path_produto': os.getenv('PATH-PRODUTO'),
        'path_clientes': os.getenv('PATH-CLIENTES'),
        'path_clientes_sp': os.getenv('PATH-CLIENTES-SP'),
        'path_clientes_pr': os.getenv('PATH-CLIENTES-PR'),
        'path_estoque': os.getenv('PATH-ESTOQUE'),
        'path_estoque_sp': os.getenv('PATH-ESTOQUE-SP'),
        'path_estoque_pr': os.getenv('PATH-ESTOQUE-PR'),
        'path_vendas': os.getenv('PATH-VENDAS'),
        'path_vendas_sp': os.getenv('PATH-VENDAS-SP'),
        'path_vendas_pr': os.getenv('PATH-VENDAS-PR'),
    }

    cod_estados = ['PR','SP']
    
    conn = get_db_engine()
    ftp = FTP_CONFIG

    for cod_estado in cod_estados:
        query = (f"""
            (
                SELECT 
                    clie.clie_unid_codigo AS unidade,
                    clie.clie_codigo AS clie_codigo,
                    clie.clie_nome AS clie_nome,
                    clie.clie_cnpjcpf AS cnpjcpf,
                    UPPER(clie.clie_razaosocial) AS razaosocial,
                    UPPER(clie.clie_endres) AS endereco,
                    UPPER(clie.clie_endresnumero) AS numero_res,
                    UPPER(clie.clie_bairrores) AS bairro,
                    clie.clie_cepres AS cep,
                    clie.clie_muni_codigo_res AS code_muni,
                    muni.muni_codigo AS cod_municipio,
                    UPPER(muni.muni_nome) AS municipio,
                    clie.clie_ufexprg AS estado,
                    clie.clie_vend_codigo AS vend_codigo1,
                    clie.clie_vend_alternativos AS vend_codigo2,
                    vend.vend_nome AS vend_nome,
                    clie.clie_dtcad,
                    (clie.clie_vend_codigo ||';'|| clie.clie_vend_alternativos) AS vend_codigo_all
                FROM clientes AS clie 
                LEFT JOIN municipios AS muni ON clie.clie_muni_codigo_res = muni.muni_codigo
                LEFT JOIN movprodc AS mprc ON clie.clie_codigo = mprc.mprc_codentidade
                LEFT JOIN vendedores AS vend ON clie.clie_vend_codigo = vend.vend_codigo
                WHERE clie.clie_tipos NOT IN ('','IN','VE','FU','UN','NL')
                AND clie.clie_ufexprg = '{cod_estado}'
                AND clie.clie_endres NOT IN ('') 
                AND muni.muni_nome NOT IN ('') 
                AND clie.clie_rota_codigo NOT IN ('') 
                AND clie.clie_cnpjcpf > '0'
                AND clie.clie_cepres NOT IN ('')
                AND clie.clie_dtcad > CURRENT_DATE - INTERVAL '7 DAYS'
                GROUP BY clie.clie_unid_codigo, clie.clie_codigo, clie.clie_nome, clie.clie_cnpjcpf, clie.clie_razaosocial, clie.clie_endres, clie.clie_endresnumero, clie.clie_bairrores, clie.clie_cepres, clie.clie_muni_codigo_res, muni.muni_codigo, muni.muni_nome, clie.clie_ufexprg, clie.clie_rota_codigo, clie.clie_ramoatividade, clie.clie_vend_codigo, clie.clie_vend_alternativos, vend.vend_nome, clie.clie_dtcad
            )  
        """)

        df = pd.read_sql_query(query, conn)

        wb = openpyxl.Workbook()
        ws = wb.active

        ws['A1'] = ('systemId')
        ws['B1'] = ('name')
        ws['C1'] = ('address.countryId')
        ws['D1'] = ('address.stateId')
        ws['E1'] = ('address.city')
        ws['F1'] = ('address.postalCode')
        ws['G1'] = ('salesman')
        ws['H1'] = ('pos.active')
        ws['I1'] = ('segmentacion_global')
        for index, row in df.iterrows():
            systemId = row["clie_codigo"]
            name = row["clie_nome"]
            countryId = 2
            stateId = ''
            city = row["municipio"]
            postalCode = row["cep"]
            salesman = row["vend_nome"]
            pos_active = 'True'
            segmentacion_global = 'OUTROS'

            ws.cell(row=index+2, column=1).value = (f'{systemId:.0f}')
            ws.cell(row=index+2, column=2).value = (f'{name}')
            ws.cell(row=index+2, column=3).value = (f'{countryId}')
            ws.cell(row=index+2, column=4).value = (f'{stateId}')
            ws.cell(row=index+2, column=5).value = (f'{city}')
            ws.cell(row=index+2, column=6).value = (f'{postalCode}')
            ws.cell(row=index+2, column=7).value = (f'{salesman}')
            ws.cell(row=index+2, column=8).value = (f'{pos_active}')
            ws.cell(row=index+2, column=9).value = (f'{segmentacion_global}')

        dataAtual = datetime.now().strftime("%Y-%m-%d")
        nomeArquivo = (f'CLIENTESDUSNEI{cod_estado}{dataAtual}')
        ws.title = dataAtual
        diretorio = f'C:/Users/Windows/Documents/Python/mccain-automation/app/data/{dataAtual}'
        if not os.path.exists(diretorio):
            os.mkdir(diretorio)
        local_arquivo = os.path.join(
            f'C:/Users/Windows/Documents/Python/mccain-automation/app/data/{dataAtual}/{nomeArquivo}.xlsx')

        wb.save(local_arquivo)


    with FTP(FTP_CONFIG['server-ftp']) as ftp:
        ftp.login(user=FTP_CONFIG['user-ftp'], passwd=FTP_CONFIG['password-ftp'])

        remote_dir_path_pr = os.path.join(FTP_CONFIG['path_clientes_pr'])
        remote_dir_path_sp = os.path.join(FTP_CONFIG['path_clientes_sp'])

        # try:
        #     ftp.mkd(remote_dir_path)
        #     print(f'Diretório {remote_dir_path} criado!')
        # except Exception as e:
        #     print('Não foi possível criar a pasta, pode ser que já exista!')

        for arquivos_data in os.listdir(diretorio):
            if 'CLIENTESDUSNEIPR' in arquivos_data:
                file_path = os.path.join(diretorio, arquivos_data)

                if os.path.isfile(file_path):
                    with open(local_arquivo, 'rb') as local_file:
                        remote_path = os.path.join(remote_dir_path_pr, arquivos_data)
                        ftp.storbinary(f"STOR {remote_path}", local_file)
                logging.info(
                    f"Arquivo {os.path.basename(arquivos_data)} upload FTP server concluído com sucesso!")
                
            if 'CLIENTESDUSNEISP' in arquivos_data:
                file_path = os.path.join(diretorio, arquivos_data)

                if os.path.isfile(file_path):
                    with open(local_arquivo, 'rb') as local_file:
                        remote_path = os.path.join(remote_dir_path_sp, arquivos_data)
                        ftp.storbinary(f"STOR {remote_path}", local_file)
                logging.info(
                    f"Arquivo {os.path.basename(arquivos_data)} upload FTP server concluído com sucesso!")
            

if __name__ == "__main__":
    clientes()