#!/usr/bin/env python3

import os
import ftplib
import openpyxl
import pandas as pd
from datetime import datetime
from config import conn_engine, ConnectionData

def vendas_001():
    conn = conn_engine(ConnectionData.DB_CONFIG)
    ftp = ConnectionData.FTP_CONFIG

    wb = openpyxl.Workbook()
    ws = wb.active
    
    query = ("""
    (
        SELECT
        mprd.mprd_dcto_codigo AS doc_cod, 
        mprd.mprd_transacao AS transacao,
        clie.clie_cnpjcpf AS cnpj_cpf,
        clie.clie_codigo AS cod_clie,
        mprd.mprd_datamvto AS data,
        mprd.mprd_numerodcto AS nfe,
        prod.prod_codbarras AS cod_barras,
        prod.prod_codigo AS cod_prod,
        (mprd.mprd_qtde * prod.prod_pesoliq) AS quantity,
        mprd.mprd_valor AS amount,
        mprc.mprc_vend_codigo AS cod_vend,
        SUBSTRING(clie.clie_cepres, 1,5) ||'-'|| SUBSTRING(clie.clie_cepres, 6,3) AS cep
        FROM movprodd0523 AS mprd 
        LEFT JOIN movprodc AS mprc ON mprd.mprd_operacao = mprc.mprc_operacao
        LEFT JOIN produtos AS prod ON mprd.mprd_prod_codigo = prod.prod_codigo
        LEFT JOIN clientes AS clie ON mprc.mprc_codentidade = clie.clie_codigo
        WHERE mprd_status = 'N' 
        AND mprd_unid_codigo IN ('001')
        AND prod.prod_marca IN ('MCCAIN','MCCAIN RETAIL')
        AND mprd.mprd_dcto_codigo IN ('6666','6668','7335','7337','7338','7339','7260','7263','7262','7268','7264','7269', '7267', '7319', '7318')
        AND mprd.mprd_datamvto > CURRENT_DATE - INTERVAL '7 DAYS'
    )  
        UNION ALL      
    (
        SELECT
        mprd.mprd_dcto_codigo AS doc_cod, 
        mprd.mprd_transacao AS transacao,
        clie.clie_cnpjcpf AS cnpj_cpf,
        clie.clie_codigo AS cod_clie,
        mprd.mprd_datamvto AS data,
        mprd.mprd_numerodcto AS nfe,
        prod.prod_codbarras AS cod_barras,
        prod.prod_codigo AS cod_prod,
        (mprd.mprd_qtde * prod.prod_pesoliq) AS quantity,
        mprd.mprd_valor AS amount,
        mprc.mprc_vend_codigo AS cod_vend,
        SUBSTRING(clie.clie_cepres, 1,5) ||'-'|| SUBSTRING(clie.clie_cepres, 6,3) AS cep
        FROM movprodd0623 AS mprd 
        LEFT JOIN movprodc AS mprc ON mprd.mprd_operacao = mprc.mprc_operacao
        LEFT JOIN produtos AS prod ON mprd.mprd_prod_codigo = prod.prod_codigo
        LEFT JOIN clientes AS clie ON mprc.mprc_codentidade = clie.clie_codigo
        WHERE mprd_status = 'N' 
        AND mprd_unid_codigo IN ('001')
        AND prod.prod_marca IN ('MCCAIN','MCCAIN RETAIL')
        AND mprd.mprd_dcto_codigo IN ('6666','6668','7335','7337','7338','7339','7260','7263','7262','7268','7264','7269', '7267', '7319', '7318')
        AND mprd.mprd_datamvto > CURRENT_DATE - INTERVAL '7 DAYS'
    )    
        UNION ALL    
    (
        SELECT
        mprd.mprd_dcto_codigo AS doc_cod, 
        mprd.mprd_transacao AS transacao,
        clie.clie_cnpjcpf AS cnpj_cpf,
        clie.clie_codigo AS cod_clie,
        mprd.mprd_datamvto AS data,
        mprd.mprd_numerodcto AS nfe,
        prod.prod_codbarras AS cod_barras,
        prod.prod_codigo AS cod_prod,
        (mprd.mprd_qtde * prod.prod_pesoliq) AS quantity,
        mprd.mprd_valor AS amount,
        mprc.mprc_vend_codigo AS cod_vend,
        SUBSTRING(clie.clie_cepres, 1,5) ||'-'|| SUBSTRING(clie.clie_cepres, 6,3) AS cep
        FROM movprodd0723 AS mprd 
        LEFT JOIN movprodc AS mprc ON mprd.mprd_operacao = mprc.mprc_operacao
        LEFT JOIN produtos AS prod ON mprd.mprd_prod_codigo = prod.prod_codigo
        LEFT JOIN clientes AS clie ON mprc.mprc_codentidade = clie.clie_codigo
        WHERE mprd_status = 'N' 
        AND mprd_unid_codigo IN ('001')
        AND prod.prod_marca IN ('MCCAIN','MCCAIN RETAIL')
        AND mprd.mprd_dcto_codigo IN ('6666','6668','7335','7337','7338','7339','7260','7263','7262','7268','7264','7269', '7267', '7319', '7318')
        AND mprd.mprd_datamvto > CURRENT_DATE - INTERVAL '7 DAYS'
    )      
        UNION ALL  
    (
        SELECT
        mprd.mprd_dcto_codigo AS doc_cod, 
        mprd.mprd_transacao AS transacao,
        clie.clie_cnpjcpf AS cnpj_cpf,
        clie.clie_codigo AS cod_clie,
        mprd.mprd_datamvto AS data,
        mprd.mprd_numerodcto AS nfe,
        prod.prod_codbarras AS cod_barras,
        prod.prod_codigo AS cod_prod,
        (mprd.mprd_qtde * prod.prod_pesoliq) AS quantity,
        mprd.mprd_valor AS amount,
        mprc.mprc_vend_codigo AS cod_vend,
        SUBSTRING(clie.clie_cepres, 1,5) ||'-'|| SUBSTRING(clie.clie_cepres, 6,3) AS cep
        FROM movprodd0823 AS mprd 
        LEFT JOIN movprodc AS mprc ON mprd.mprd_operacao = mprc.mprc_operacao
        LEFT JOIN produtos AS prod ON mprd.mprd_prod_codigo = prod.prod_codigo
        LEFT JOIN clientes AS clie ON mprc.mprc_codentidade = clie.clie_codigo
        WHERE mprd_status = 'N' 
        AND mprd_unid_codigo IN ('001')
        AND prod.prod_marca IN ('MCCAIN','MCCAIN RETAIL')
        AND mprd.mprd_dcto_codigo IN ('6666','6668','7335','7337','7338','7339','7260','7263','7262','7268','7264','7269', '7267', '7319', '7318')
        AND mprd.mprd_datamvto > CURRENT_DATE - INTERVAL '7 DAYS'
    )        
        UNION ALL  
    (
        SELECT
        mprd.mprd_dcto_codigo AS doc_cod, 
        mprd.mprd_transacao AS transacao,
        clie.clie_cnpjcpf AS cnpj_cpf,
        clie.clie_codigo AS cod_clie,
        mprd.mprd_datamvto AS data,
        mprd.mprd_numerodcto AS nfe,
        prod.prod_codbarras AS cod_barras,
        prod.prod_codigo AS cod_prod,
        (mprd.mprd_qtde * prod.prod_pesoliq) AS quantity,
        mprd.mprd_valor AS amount,
        mprc.mprc_vend_codigo AS cod_vend,
        SUBSTRING(clie.clie_cepres, 1,5) ||'-'|| SUBSTRING(clie.clie_cepres, 6,3) AS cep
        FROM movprodd0923 AS mprd 
        LEFT JOIN movprodc AS mprc ON mprd.mprd_operacao = mprc.mprc_operacao
        LEFT JOIN produtos AS prod ON mprd.mprd_prod_codigo = prod.prod_codigo
        LEFT JOIN clientes AS clie ON mprc.mprc_codentidade = clie.clie_codigo
        WHERE mprd_status = 'N' 
        AND mprd_unid_codigo IN ('001')
        AND prod.prod_marca IN ('MCCAIN','MCCAIN RETAIL')
        AND mprd.mprd_dcto_codigo IN ('6666','6668','7335','7337','7338','7339','7260','7263','7262','7268','7264','7269', '7267', '7319', '7318')
        AND mprd.mprd_datamvto > CURRENT_DATE - INTERVAL '7 DAYS'
    )        
        UNION ALL  
    (
        SELECT
        mprd.mprd_dcto_codigo AS doc_cod, 
        mprd.mprd_transacao AS transacao,
        clie.clie_cnpjcpf AS cnpj_cpf,
        clie.clie_codigo AS cod_clie,
        mprd.mprd_datamvto AS data,
        mprd.mprd_numerodcto AS nfe,
        prod.prod_codbarras AS cod_barras,
        prod.prod_codigo AS cod_prod,
        (mprd.mprd_qtde * prod.prod_pesoliq) AS quantity,
        mprd.mprd_valor AS amount,
        mprc.mprc_vend_codigo AS cod_vend,
        SUBSTRING(clie.clie_cepres, 1,5) ||'-'|| SUBSTRING(clie.clie_cepres, 6,3) AS cep
        FROM movprodd1023 AS mprd 
        LEFT JOIN movprodc AS mprc ON mprd.mprd_operacao = mprc.mprc_operacao
        LEFT JOIN produtos AS prod ON mprd.mprd_prod_codigo = prod.prod_codigo
        LEFT JOIN clientes AS clie ON mprc.mprc_codentidade = clie.clie_codigo
        WHERE mprd_status = 'N' 
        AND mprd_unid_codigo IN ('001')
        AND prod.prod_marca IN ('MCCAIN','MCCAIN RETAIL')
        AND mprd.mprd_dcto_codigo IN ('6666','6668','7335','7337','7338','7339','7260','7263','7262','7268','7264','7269', '7267', '7319', '7318')
        AND mprd.mprd_datamvto > CURRENT_DATE - INTERVAL '7 DAYS'
    )        
        UNION ALL  
    (
        SELECT
        mprd.mprd_dcto_codigo AS doc_cod, 
        mprd.mprd_transacao AS transacao,
        clie.clie_cnpjcpf AS cnpj_cpf,
        clie.clie_codigo AS cod_clie,
        mprd.mprd_datamvto AS data,
        mprd.mprd_numerodcto AS nfe,
        prod.prod_codbarras AS cod_barras,
        prod.prod_codigo AS cod_prod,
        (mprd.mprd_qtde * prod.prod_pesoliq) AS quantity,
        mprd.mprd_valor AS amount,
        mprc.mprc_vend_codigo AS cod_vend,
        SUBSTRING(clie.clie_cepres, 1,5) ||'-'|| SUBSTRING(clie.clie_cepres, 6,3) AS cep
        FROM movprodd1123 AS mprd 
        LEFT JOIN movprodc AS mprc ON mprd.mprd_operacao = mprc.mprc_operacao
        LEFT JOIN produtos AS prod ON mprd.mprd_prod_codigo = prod.prod_codigo
        LEFT JOIN clientes AS clie ON mprc.mprc_codentidade = clie.clie_codigo
        WHERE mprd_status = 'N' 
        AND mprd_unid_codigo IN ('001')
        AND prod.prod_marca IN ('MCCAIN','MCCAIN RETAIL')
        AND mprd.mprd_dcto_codigo IN ('6666','6668','7335','7337','7338','7339','7260','7263','7262','7268','7264','7269', '7267', '7319', '7318')
        AND mprd.mprd_datamvto > CURRENT_DATE - INTERVAL '7 DAYS'
    )        
        UNION ALL  
    (
        SELECT
        mprd.mprd_dcto_codigo AS doc_cod, 
        mprd.mprd_transacao AS transacao,
        clie.clie_cnpjcpf AS cnpj_cpf,
        clie.clie_codigo AS cod_clie,
        mprd.mprd_datamvto AS data,
        mprd.mprd_numerodcto AS nfe,
        prod.prod_codbarras AS cod_barras,
        prod.prod_codigo AS cod_prod,
        (mprd.mprd_qtde * prod.prod_pesoliq) AS quantity,
        mprd.mprd_valor AS amount,
        mprc.mprc_vend_codigo AS cod_vend,
        SUBSTRING(clie.clie_cepres, 1,5) ||'-'|| SUBSTRING(clie.clie_cepres, 6,3) AS cep
        FROM movprodd1223 AS mprd 
        LEFT JOIN movprodc AS mprc ON mprd.mprd_operacao = mprc.mprc_operacao
        LEFT JOIN produtos AS prod ON mprd.mprd_prod_codigo = prod.prod_codigo
        LEFT JOIN clientes AS clie ON mprc.mprc_codentidade = clie.clie_codigo
        WHERE mprd_status = 'N' 
        AND mprd_unid_codigo IN ('001')
        AND prod.prod_marca IN ('MCCAIN','MCCAIN RETAIL')
        AND mprd.mprd_dcto_codigo IN ('6666','6668','7335','7337','7338','7339','7260','7263','7262','7268','7264','7269', '7267', '7319', '7318')
        AND mprd.mprd_datamvto > CURRENT_DATE - INTERVAL '7 DAYS'
    )        
    """)
    df = pd.read_sql_query(query, conn)
    
    
    ws['A1'] = (f'systemId')
    ws['B1'] = (f'Code')
    ws['C1'] = (f'Quantity')
    ws['D1'] = (f'Amount')
    ws['E1'] = (f'Sale Date')
    ws['F1'] = (f'Transaction ID')
    ws['G1'] = (f'DocCod')
    for i in range(1, len(df)):
        systemId = df.loc[i, "cod_clie"]
        code = df.loc[i, "cod_prod"]
        doc_cod = df.loc[i, "doc_cod"]
        quantity = df.loc[i, "quantity"]
        
        if doc_cod == '7260' or doc_cod == '7263' or doc_cod == '7262' or doc_cod == '7268' or doc_cod == '7264' or doc_cod == '7269' or doc_cod == '7267' or doc_cod == '7319' or doc_cod == '7318':
            quantity = -quantity
        
        amount = str(df.loc[i, "amount"]).replace(',','.')
        amount2 = float(amount)
        data = df.loc[i, "data"].strftime("%Y-%m-%d")
        transactionId = "1" + df.loc[i, "nfe"].zfill(6)
        
        ws.cell(row=i+1, column = 1).value = (f'{systemId:.0f}')
        ws.cell(row=i+1, column = 2).value = (f'{code:.0f}')
        ws.cell(row=i+1, column = 3).value = (f'{quantity:.2f}')
        ws.cell(row=i+1, column = 4).value = (f'{amount2:.2f}')
        ws.cell(row=i+1, column = 5).value = (f'{data}')
        ws.cell(row=i+1, column = 6).value = (f'{transactionId}')
        ws.cell(row=i+1, column = 7).value = (f'{doc_cod}')
        
    dataAtual = datetime.now().strftime("%Y-%m-%d")
    nomeArquivo = (f'VENDASDUSNEI001{dataAtual}')
    ws.title = dataAtual
    local_arquivo = os.path.join(f'C:/Users/Windows/Documents/Python/mccain-automation/app/app/data/{dataAtual}/{nomeArquivo}.xlsx')
    wb.save(local_arquivo)
    
    
    with ftplib.FTP(ftp['server-ftp']) as ftp:
        ftp.login(user=ftp['user-ftp'], passwd=ftp['password-ftp'])
        with open(local_arquivo, 'rb') as local_file:
            remote_path = os.path.join(ftp['path_clientes'], os.path.basename(local_arquivo))
            ftp.storbinary(f"STOR {remote_path}", local_file)
    print(f"Arquivo {os.path.basename(local_arquivo)} upload FTP server concluído com sucesso!")

if __name__ == "__main__":
  vendas_001()
  