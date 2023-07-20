import os
import logging
import openpyxl
import pandas as pd
from ftplib import FTP
from datetime import datetime
from dotenv import load_dotenv
from conn import get_db_engine

load_dotenv()

def vendas():
    FTP_CONFIG = {
        'server-ftp': os.getenv('SERVER-FTP'),
        'user-ftp': os.getenv('USER-FTP'),
        'password-ftp': os.getenv('PASSWORD-FTP'),
        'path_clientes': os.getenv('PATH-CLIENTES'),
        'path_estoque': os.getenv('PATH-ESTOQUE'),
        'path_produto': os.getenv('PATH-PRODUTO'),
        'path_vendas': os.getenv('PATH-VENDAS'),
    }

    unid_codigos = ['001', '002', '003']

    for unid_codigo in unid_codigos:

        def vendas_query(table_name, conn, unid_codigo):
            query = (f"""
                (
                    SELECT
                    mprd.mprd_dcto_codigo AS doc_cod, 
                    mprd.mprd_transacao AS transacao,
                    clie.clie_cnpjcpf AS cnpj_cpf,
                    clie.clie_codigo AS cod_clie,
                    mprd.mprd_datamvto AS data,
                    mprd.mprd_numerodcto AS nfe,
                    prod.prod_codbarras AS cod_barras,
                    prod.prod_codigo AS cod_prod,
                    (mprd.mprd_qtde * prod.prod_pesoliq) AS quantity,
                    mprd.mprd_valor AS amount,
                    mprc.mprc_vend_codigo AS cod_vend,
                    SUBSTRING(clie.clie_cepres, 1,5) ||'-'|| SUBSTRING(clie.clie_cepres, 6,3) AS cep
                    FROM {table_name} AS mprd 
                    LEFT JOIN movprodc AS mprc ON mprd.mprd_operacao = mprc.mprc_operacao
                    LEFT JOIN produtos AS prod ON mprd.mprd_prod_codigo = prod.prod_codigo
                    LEFT JOIN clientes AS clie ON mprc.mprc_codentidade = clie.clie_codigo
                    WHERE mprd_status = 'N' 
                    AND mprd_unid_codigo IN ('{unid_codigo}')
                    AND prod.prod_marca IN ('MCCAIN','MCCAIN RETAIL')
                    AND mprd.mprd_dcto_codigo IN ('6666','6668','7335','7337','7338','7339','7260','7263','7262','7268','7264','7269', '7267', '7319', '7318')
                    AND mprd.mprd_datamvto > CURRENT_DATE - INTERVAL '7 DAYS'
                )  
            """)
            return pd.read_sql_query(query, conn)

        conn = get_db_engine()
        ftp = FTP_CONFIG

        wb = openpyxl.Workbook()
        ws = wb.active

        tables = ['movprodd0523', 'movprodd0623', 'movprodd0723', 'movprodd0823',
                'movprodd0923', 'movprodd1023', 'movprodd1123', 'movprodd1223']

        df = pd.concat([vendas_query(table, conn, unid_codigo)for table in tables])

        ws['A1'] = (f'systemId')
        ws['B1'] = (f'Code')
        ws['C1'] = (f'Quantity')
        ws['D1'] = (f'Amount')
        ws['E1'] = (f'Sale Date')
        ws['F1'] = (f'Transaction ID')
        for index, row in df.iterrows():
            systemId = row["cod_clie"]
            code = row["cod_prod"]
            doc_cod = row["doc_cod"]
            quantity = row["quantity"]

            if doc_cod in ['7260', '7263', '7262', '7268', '7264', '7269', '7267', '7319', '7318']:
                quantity = -quantity

            amount = str(row["amount"]).replace(',', '.')
            amount2 = float(amount)
            data = row["data"].strftime("%Y-%m-%d")
            transactionId = "1" + row["nfe"].zfill(6)

            ws.cell(row=index+2, column=1).value = (f'{systemId:.0f}')
            ws.cell(row=index+2, column=2).value = (f'{code:.0f}')
            ws.cell(row=index+2, column=3).value = (f'{quantity:.2f}')
            ws.cell(row=index+2, column=4).value = (f'{amount2:.2f}')
            ws.cell(row=index+2, column=5).value = (f'{data}')
            ws.cell(row=index+2, column=6).value = (f'{transactionId}')

        dataAtual = datetime.now().strftime("%Y-%m-%d")
        nomeArquivo = (f'VENDASDUSNEI{unid_codigo}{dataAtual}')
        ws.title = dataAtual
        diretorio = f'C:/Users/Windows/Documents/Python/mccain-automation/app/data/{dataAtual}'
        if not os.path.exists(diretorio):
            os.mkdir(diretorio, exist_ok=True)
        local_arquivo = os.path.join(
            f'C:/Users/Windows/Documents/Python/mccain-automation/app/data/{dataAtual}/{nomeArquivo}.xlsx')

        wb.save(local_arquivo)


    # with FTP(FTP_CONFIG['server-ftp']) as ftp:
    #     ftp.login(user=FTP_CONFIG['user-ftp'], passwd=FTP_CONFIG['password-ftp'])

    #     remote_dir_path = os.path.join(FTP_CONFIG['path_vendas'])

    #     for arquivos_data in os.listdir(diretorio):
    #         if 'VENDAS' in arquivos_data:
    #             file_path = os.path.join(diretorio, arquivos_data)

    #             if os.path.isfile(file_path):
    #                 with open(local_arquivo, 'rb') as local_file:
    #                     remote_path = os.path.join(remote_dir_path, arquivos_data)
    #                     ftp.storbinary(f"STOR {remote_path}", local_file)
    #             logging.info(
    #                 f"Arquivo {os.path.basename(arquivos_data)} upload FTP server concluído com sucesso!")

if __name__ == "__main__":
  vendas()
  